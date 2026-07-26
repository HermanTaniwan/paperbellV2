import json
import sys
from datetime import date, datetime, time
from openpyxl import load_workbook
from openpyxl.utils.datetime import to_excel

if len(sys.argv) != 3:
    raise SystemExit("usage: xlsx_to_json.py input.xlsx output.json")

workbook = load_workbook(sys.argv[1], read_only=True, data_only=True)
sheet = workbook.worksheets[0]
rows = []
for row in sheet.iter_rows(values_only=True):
    converted = []
    for value in row:
        if value is None:
            converted.append("")
        elif isinstance(value, (datetime, date, time)):
            converted.append(to_excel(value, workbook.epoch))
        else:
            converted.append(value)
    rows.append(converted)
with open(sys.argv[2], "w", encoding="utf-8") as handle:
    json.dump(rows, handle, ensure_ascii=False)
